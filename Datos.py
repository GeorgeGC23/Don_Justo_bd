import sqlite3
from datetime import datetime
from openpyxl import Workbook
# ... (Código de inicialización de la base de datos y menú principal) ...
# Función para crear la base de datos SQLite y tablas si no existen
def inicializar_base_de_datos():
    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio REAL NOT NULL,
            fecha DATE NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS compras (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            proveedor TEXT NOT NULL,
            producto TEXT NOT NULL,
            cantidad INTEGER NOT NULL,
            precio REAL NOT NULL,
            fecha DATE NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS almacen (
            producto TEXT PRIMARY KEY,
            cantidad INTEGER NOT NULL
        )
    ''')
    conn.commit()
    conn.close()

# Función para mostrar el menú principal y procesar las opciones del usuario
def menu_principal():
    print("=== Menú Principal ===")
    print("1. Registrar una Venta")
    print("2. Ver Reporte de Ventas")
    print("3. Registrar una Compra")
    print("4. Ver Reporte de Compras")
    print("5. Ver Almacen")
    print("6. Salir")

    opcion = input("Ingrese el número de la opción deseada: ")

    if opcion == '1':
        registrar_venta()
    elif opcion == '2':
        ver_reporte_ventas()
    elif opcion == '3':
        registrar_compra()
    elif opcion == '4':
        ver_reporte_compras()
    elif opcion == '5':
        ver_almacen()
    elif opcion == '6':
        print("¡Adiós!")
        exit()
    else:
        print("Opción no válida. Por favor, ingrese un número válido.")
    
    # Después de procesar una opción, volvemos al menú principal
    menu_principal()
    
    # Función para ver el estado actual del almacen
def ver_almacen():
    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM almacen')
    almacen = cursor.fetchall()

    conn.close()

    if not almacen:
        print("El almacen está vacío.")
    else:
        print("=== Almacen ===")
        for producto in almacen:
            print(f"Producto: {producto[0]}, Cantidad: {producto[1]}")
    
# Función para registrar una venta
def registrar_venta():
    producto = input("Ingrese el nombre del producto vendido: ")
    cantidad = int(input("Ingrese la cantidad vendida: "))
    precio = float(input("Ingrese el precio unitario: "))
    fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

 # Verificar la disponibilidad del producto en el almacen
    cursor.execute('SELECT cantidad FROM almacen WHERE producto = ?', (producto,))
    almacen_info = cursor.fetchone()

    if not almacen_info or almacen_info[0] < cantidad:
        print(f"El producto '{producto}' no está disponible en cantidad suficiente en el almacen.")
        conn.close()
        return

    # Registrar la venta
    cursor.execute('''
        INSERT INTO ventas (producto, cantidad, precio, fecha)
        VALUES (?, ?, ?, ?)
    ''', (producto, cantidad, precio, fecha))
    
     # Actualizar el registro del almacen restando la cantidad vendida
    cursor.execute('''
        UPDATE almacen
        SET cantidad = cantidad - ?
        WHERE producto = ?
    ''', (cantidad, producto))

# Exportar los datos actualizados del almacen y las ventas a los archivos CSV
    exportar_almacen_a_xlsx()
    exportar_ventas_a_xlsx()

    conn.commit()
    conn.close()

    print("Venta registrada correctamente.")

# Función para ver el reporte de ventas
def ver_reporte_ventas():
    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM ventas')
    ventas = cursor.fetchall()

    conn.close()

    if not ventas:
        print("No hay ventas registradas.")
    else:
        print("=== Reporte de Ventas ===")
        for venta in ventas:
            print(f"ID: {venta[0]}, Producto: {venta[1]}, Cantidad: {venta[2]}, Precio: {venta[3]}, Fecha: {venta[4]}")

# Función para registrar una compra
def registrar_compra():
    proveedor = input("Ingrese el nombre del proveedor: ")
    producto = input("Ingrese el nombre del producto comprado: ")
    cantidad = int(input("Ingrese la cantidad comprada: "))
    precio = float(input("Ingrese el precio unitario: "))
    fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    # Registrar la compra en la tabla 'compras'
    cursor.execute('''
        INSERT INTO compras (proveedor, producto, cantidad, precio, fecha)
        VALUES (?, ?, ?, ?, ?)
    ''', (proveedor, producto, cantidad, precio, fecha))
    
    # Actualizar el registro del almacen o crearlo si no existe
    cursor.execute('''
        INSERT INTO almacen (producto, cantidad)
        VALUES (?, ?)
        ON CONFLICT(producto) DO UPDATE SET cantidad = cantidad + excluded.cantidad
    ''', (producto, cantidad))
    
    conn.commit()
    conn.close()

 # Exportar los datos actualizados del almacen y las compras a los archivos CSV
    exportar_almacen_a_xlsx()
    exportar_compras_a_xlsx()

    print("Compra registrada correctamente.")

# Función para ver el reporte de compras
def ver_reporte_compras():
    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM compras')
    compras = cursor.fetchall()

    conn.close()

    if not compras:
        print("No hay compras registradas.")
    else:
        print("=== Reporte de Compras ===")
        for compra in compras:
            print(f"ID: {compra[0]}, Proveedor: {compra[1]}, Producto: {compra[2]}, Cantidad: {compra[3]}, Precio: {compra[4]}, Fecha: {compra[5]}")

# Función para exportar los datos de la tabla 'almacen' a un archivo CSV
# Función para exportar los datos de la tabla 'almacen' a un archivo XLSX
def exportar_almacen_a_xlsx():
    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM almacen')
    almacen = cursor.fetchall()

    conn.close()

    if not almacen:
        print("No hay datos en el almacen para exportar.")
        return

    # Crear un nuevo libro de Excel
    workbook = Workbook()
    sheet = workbook.active

    # Agregar encabezados de columna
    sheet.append(['Producto', 'Cantidad'])

    # Agregar datos
    for row in almacen:
        sheet.append(row)

    # Guardar el libro de Excel como archivo XLSX
    workbook.save('almacen.xlsx')

    print("Datos del almacen exportados a almacen.xlsx")

# Función para exportar los datos de la tabla 'ventas' a un archivo XLSX
def exportar_ventas_a_xlsx():
    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM ventas')
    ventas = cursor.fetchall()

    conn.close()

    if not ventas:
        print("No hay datos de ventas para exportar.")
        return

    # Crear un nuevo libro de Excel
    workbook = Workbook()
    sheet = workbook.active

    # Agregar encabezados de columna
    sheet.append(['ID', 'Producto', 'Cantidad', 'Precio', 'Fecha'])

    # Agregar datos
    for row in ventas:
        sheet.append(row)

    # Guardar el libro de Excel como archivo XLSX
    workbook.save('ventas.xlsx')

    print("Datos de ventas exportados a ventas.xlsx")

# Función para exportar los datos de la tabla 'compras' a un archivo XLSX
def exportar_compras_a_xlsx():
    conn = sqlite3.connect('ventas.db')
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM compras')
    compras = cursor.fetchall()

    conn.close()

    if not compras:
        print("No hay datos de compras para exportar.")
        return

    # Crear un nuevo libro de Excel
    workbook = Workbook()
    sheet = workbook.active

    # Agregar encabezados de columna
    sheet.append(['ID', 'Proveedor', 'Producto', 'Cantidad', 'Precio', 'Fecha'])

    # Agregar datos
    for row in compras:
        sheet.append(row)

    # Guardar el libro de Excel como archivo XLSX
    workbook.save('compras.xlsx')

    print("Datos de compras exportados a compras.xlsx")
# ... (Resto del código, incluyendo el menú principal) ...

if __name__ == '__main__':
    inicializar_base_de_datos()
    menu_principal()
