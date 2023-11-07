from recursos.base_de_datos import BaseDeDatos  # Importa la clase BaseDeDatos desde el archivo base_de_datos.py

import os
import sqlite3
from openpyxl.styles import Font
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook

class LogicaDeNegocios:
    def __init__(self):
        self.base_de_datos = BaseDeDatos()  # Instancia la clase BaseDeDatos

    def confirmar_venta(self, producto, cantidad, precio_unitario, unidad, proveedor):
        conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT id_producto, producto, cantidad, unidad, proveedor, precio_unitario FROM almacen')
        productos_disponibles = cursor.fetchall()

        # Verificar si el producto está disponible en el almacén
        producto_encontrado = None
        for id_producto, nombre, cantidad_disponible, unidad_almacen, proveedor_almacen, precio in productos_disponibles:
            if nombre == producto and unidad_almacen == unidad and proveedor_almacen == proveedor:
                producto_encontrado = {
                    "id_producto": id_producto,
                    "cantidad_disponible": cantidad_disponible,
                    "precio_unitario": precio
                }
                break

        if producto_encontrado:
            id_producto = producto_encontrado["id_producto"]
            cantidad_disponible = producto_encontrado["cantidad_disponible"]
            precio_unitario = producto_encontrado["precio_unitario"]

            if cantidad <= cantidad_disponible:
                subtotal = cantidad * precio_unitario  # Calcular el subtotal de la venta

                nuevo_id_ventas = self.base_de_datos.generar_id_unico_ventas()  # Generar un nuevo ID de venta
                cursor.execute('''
                    INSERT INTO ventas (id, producto, cantidad, precio_unitario, unidad, fecha)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (nuevo_id_ventas, producto, cantidad, precio_unitario, unidad, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

                cursor.execute('''
                    UPDATE almacen
                    SET cantidad = cantidad - ?
                    WHERE id_producto = ?
                ''', (cantidad, id_producto))

                conn.commit()
                return True
            else:
                return False
        else:
            return False

    def obtener_productos_disponibles(self):
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT id_producto, producto, cantidad, unidad, proveedor, precio_unitario FROM almacen')
        productos_disponibles = cursor.fetchall()

        return productos_disponibles

    def buscar_venta_por_id(self, id_venta):
        cursor = self.base_de_datos.cursor
        cursor.execute('SELECT * FROM ventas WHERE id = ?', (id_venta,))
        venta = cursor.fetchone()
        return venta

    # Asegúrate de haber definido correctamente la estructura de ventas_realizadas para que coincida con la base de datos. 
    def obtener_todas_las_ventas(self):
        # Consultar todas las ventas
        cursor = self.base_de_datos.cursor
        cursor.execute('SELECT * FROM ventas')
        ventas = cursor.fetchall()
        return ventas

    def obtener_datos_almacen(self):
        # Consultar todos los datos del almacén
        cursor = self.base_de_datos.cursor
        cursor.execute('SELECT id_producto, producto, cantidad, unidad, precio_unitario, proveedor FROM almacen')
        datos_almacen = cursor.fetchall()

        return datos_almacen
    
    def guardar_compra(self, proveedor, producto, cantidad, precio, unidad):
        try:
            conn = self.base_de_datos.conn
            cursor = self.base_de_datos.cursor

            # Genera un ID único para la compra
            nuevo_id_compra = self.base_de_datos.generar_id_unico()

            # Obtiene la fecha y hora actual
            fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            # Inserta la compra en la tabla 'compras'
            cursor.execute('''
                INSERT INTO compras (id, proveedor, producto, cantidad, precio_unitario, unidad, fecha)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (nuevo_id_compra, proveedor, producto, cantidad, precio, unidad, fecha))

            # Inserta o actualiza la compra en la tabla 'almacen'
            cursor.execute('''
                INSERT OR REPLACE INTO almacen (id_producto, producto, cantidad, unidad, precio_unitario, proveedor)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (nuevo_id_compra, producto, cantidad, unidad, precio, proveedor))

            conn.commit()

            self.exportar_almacen_a_xlsx()
            self.exportar_compras_a_xlsx()

            print("Compra registrada correctamente.")
        except sqlite3.Error as e:
            print(f"Error al registrar la compra: {e}")      
            
    def ver_reporte_compras(self):
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT * FROM compras')
        compras = cursor.fetchall()

        return compras         
    
    def ver_reporte_compras_almacen(self):
        cursor = self.base_de_datos.cursor

        try:
            # Consulta para obtener el reporte de compras para el almacén
            cursor.execute('''
                SELECT producto, unidad, precio_unitario, SUM(cantidad) AS total_cantidad, SUM(cantidad * precio_unitario) AS total_costo 
                FROM compras 
                GROUP BY producto, unidad
            ''')
            
            compras_por_producto = cursor.fetchall()

            if compras_por_producto:
                return compras_por_producto
            else:
                # En caso de que no haya datos, retorna una lista vacía
                return []

        except sqlite3.Error as e:
            print("Error al obtener el reporte de compras para el almacén:", str(e))
            return None

    def exportar_ventas_a_xlsx(self):
        fecha_actual = datetime.now()
        fecha_actual_str = fecha_actual.strftime('%Y-%m-%d')
        mes_actual = fecha_actual.strftime('%Y-%m')

        # Crear un rango de fechas para el mes actual
        fecha_inicial = datetime(fecha_actual.year, fecha_actual.month, 1)
        fecha_final = fecha_inicial + timedelta(days=31)  # Suponemos un máximo de 31 días en un mes
        fechas_mes_actual = [(fecha_inicial + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(31)]

        # Comprobar si ya existe un libro de Excel para el mes actual
        archivo_xlsx = f'ventas_{mes_actual}.xlsx'

        if os.path.exists(archivo_xlsx):
            # Si ya existe un libro para el mes, cargarlo
            workbook = load_workbook(filename=archivo_xlsx)
        else:
            # Si no existe, crear un nuevo libro de Excel
            workbook = Workbook()

        # Recorremos las fechas y hojas de cálculo de todos los días del mes
        for fecha in fechas_mes_actual:
            cursor = self.base_de_datos.cursor
            cursor.execute('SELECT producto, cantidad, precio_unitario FROM ventas WHERE fecha LIKE ?', (f'{fecha}%',))
            ventas_del_dia = cursor.fetchall()

            if ventas_del_dia:
                # Agregar una hoja solo si hubo ventas en ese día
                if fecha not in workbook.sheetnames:
                    workbook.create_sheet(title=fecha)

                sheet = workbook[fecha]

                # Agregar encabezados a la hoja de cálculo
                sheet['A1'] = 'Producto'
                sheet['B1'] = 'Cantidad'
                sheet['C1'] = 'Precio Unitario'
                sheet['D1'] = 'Subtotal'

                total_ventas_del_dia = 0

                for i, (producto, cantidad, precio_unitario) in enumerate(ventas_del_dia, 2):
                    sheet[f'A{i}'] = producto
                    sheet[f'B{i}'] = cantidad
                    sheet[f'C{i}'] = precio_unitario
                    subtotal = cantidad * precio_unitario
                    sheet[f'D{i}'] = subtotal
                    total_ventas_del_dia += subtotal

                # Calcular y agregar la suma total de ventas del día
                total_row = len(ventas_del_dia) + 3
                sheet[f'A{total_row}'] = 'Total del Día:'
                sheet[f'D{total_row}'] = total_ventas_del_dia

                # Establecer el formato de la celda para el total
                total_cell = sheet[f'D{total_row}']
                total_cell.font = Font(bold=True)

        # Guardar el archivo Excel
        workbook.save(archivo_xlsx)

        print(f"Reporte de ventas del mes {mes_actual} exportado a {archivo_xlsx}")

    def actualizar_precio_producto(self, id_producto, nuevo_precio):
        try:
            # Verificar si el producto existe en la base de datos
            cursor = self.base_de_datos.cursor
            cursor.execute('SELECT producto FROM almacen WHERE id_producto = ?', (id_producto,))
            producto = cursor.fetchone()

            if producto:
                # Actualizar el precio unitario en la tabla 'almacen'
                cursor.execute('''
                    UPDATE almacen
                    SET precio_unitario = ?
                    WHERE id_producto = ?
                ''', (nuevo_precio, id_producto))

                self.base_de_datos.conn.commit()
                return True  # Indicar que la actualización fue exitosa
            else:
                return False  # Indicar que el producto no fue encontrado en la base de datos
        except sqlite3.Error as e:
            print(f"Error al actualizar el precio del producto: {e}")
            return False  # Indicar que hubo un error al actualizar el precio   
 

    def limpiar_contenido_tablas(self):
        try:
            if os.path.exists('ventas.xlsx'):
                os.remove('ventas.xlsx')
            if os.path.exists('compras.xlsx'):
                os.remove('compras.xlsx')
            if os.path.exists('almacen.xlsx'):
                os.remove('almacen.xlsx')
            conn = self.base_de_datos.conn
            cursor = self.base_de_datos.cursor

            cursor.execute('DELETE FROM ventas')
            cursor.execute('DELETE FROM compras')
            cursor.execute('DELETE FROM almacen')

            conn.commit()

            print("El contenido de las tablas 'ventas', 'compras' y 'almacen' y los archivos Excel han sido limpiados.")
        except Exception as e:
            print(f"Ocurrió un error al limpiar el contenido de las tablas: {str(e)}")

    def exportar_almacen_a_xlsx(self):
       # conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT * FROM almacen')
        almacen = cursor.fetchall()

        if not almacen:
            print("No hay datos en el almacén para exportar.")
            return

        # Crear un nuevo libro de Excel
        workbook = Workbook()
        sheet = workbook.active

        # Agregar encabezados de columna
        sheet.append(['id_producto', 'Producto', 'Cantidad', 'Unidad', 'Precio', 'Proveedor'])

        # Agregar datos
        for row in almacen:
            sheet.append(row)

        # Guardar el libro de Excel como archivo XLSX
        workbook.save('almacen.xlsx')

        print("Datos del almacén exportados a almacen.xlsx")

    def exportar_compras_a_xlsx(self):
        fecha_actual = datetime.now()
        fecha_actual_str = fecha_actual.strftime('%Y-%m-%d')
        mes_actual = fecha_actual.strftime('%Y-%m')

        # Crear un rango de fechas para el mes actual
        fecha_inicial = datetime(fecha_actual.year, fecha_actual.month, 1)
        fecha_final = fecha_inicial + timedelta(days=31)  # Suponemos un máximo de 31 días en un mes
        fechas_mes_actual = [(fecha_inicial + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(31)]

        # Comprobar si ya existe un libro de Excel para el mes actual
        archivo_xlsx = f'compras_{mes_actual}.xlsx'

        if os.path.exists(archivo_xlsx):
            # Si ya existe un libro para el mes, cargarlo
            workbook = load_workbook(filename=archivo_xlsx)
        else:
            # Si no existe, crear un nuevo libro de Excel
            workbook = Workbook()

        # Crear una hoja por mes
        if mes_actual not in workbook.sheetnames:
            workbook.create_sheet(title=mes_actual)

        sheet = workbook[mes_actual]

        # Agregar encabezados a la hoja de cálculo
        sheet['A1'] = 'Producto'
        sheet['B1'] = 'Cantidad'
        sheet['C1'] = 'Precio Unitario'
        sheet['D1'] = 'Subtotal'

        total_compras_del_mes = 0
        i = 2  # Inicializar i en 2

        for fecha in fechas_mes_actual:
            cursor = self.base_de_datos.cursor
            cursor.execute('SELECT producto, cantidad, precio_unitario FROM compras WHERE fecha LIKE ?', (f'{fecha}%',))
            compras_del_dia = cursor.fetchall()

            for producto, cantidad, precio_unitario in compras_del_dia:
                sheet[f'A{i}'] = producto
                sheet[f'B{i}'] = cantidad
                sheet[f'C{i}'] = precio_unitario
                subtotal = cantidad * precio_unitario
                sheet[f'D{i}'] = subtotal
                total_compras_del_mes += subtotal

        # Calcular y agregar la suma total de compras del mes
        total_row = len(fechas_mes_actual) + 3
        sheet[f'A{total_row}'] = 'Total del Mes:'
        sheet[f'D{total_row}'] = total_compras_del_mes

        # Establecer el formato de la celda para el total
        total_cell = sheet[f'D{total_row}']
        total_cell.font = Font(bold=True)

        # Guardar el archivo Excel
        workbook.save(archivo_xlsx)

        print(f"Reporte de compras del mes {mes_actual} exportado a {archivo_xlsx}") 

if __name__ == '__main__':
    logica = LogicaDeNegocios()
    logica.base_de_datos.inicializar_base_de_datos()