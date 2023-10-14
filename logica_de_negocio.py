from base_de_datos import BaseDeDatos  # Importa la clase BaseDeDatos desde el archivo basededatos.py

import os
import openpyxl
import sqlite3

from openpyxl.styles import Font
from datetime import datetime
from openpyxl import Workbook


class LogicaDeNegocios:
    def __init__(self):
        self.base_de_datos = BaseDeDatos()  # Instancia la clase BaseDeDatos

    def menu_principal(self):
      while True:
        print("=== Menú Principal ===")
        print("1. Registrar una Venta")
        print("2. Ver Reporte de Ventas")
        print("3. Registrar una Compra")
        print("4. Ver Reporte de Compras")
        print("5. Ver Reporte de Compras para el Almacén")
        print("6. Ver Almacén")
        print("7. Limpiar Contenido de Tablas")
        print("8. Modificar Precio de un Producto")
        print("9. Salir")

        opcion = input("Ingrese el número de la opción deseada: ")

        if opcion == '1':
            self.registrar_venta()
        elif opcion == '2':
            self.ver_reporte_ventas()
        elif opcion == '3':
            self.registrar_compra()
        elif opcion == '4':
            self.ver_reporte_compras()
        elif opcion == '5':
            self.ver_reporte_compras_almacen()
        elif opcion == '6':
            self.ver_almacen()
        elif opcion == '7':
            self.limpiar_contenido_tablas()
        elif opcion == '8':
            self.modificar_precio_producto()  # Nueva opción para modificar precio
        elif opcion == '9':
            print("¡Adiós!")
            exit()
        else:
            print("Opción no válida. Por favor, ingrese un número válido.")

        # Después de procesar una opción, volvemos al menú principal
        self.menu_principal()

    def ver_almacen(self):
      #  conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT * FROM almacen')
        almacen = cursor.fetchall()

        if not almacen:
            print("El almacén está vacío.")
        else:
            print("=== Almacén ===")
            for producto in almacen:
                print(f"ID: {producto[0]}, Producto: {producto[1]}, Cantidad: {producto[2]}, Unidad: {producto[3]}, Precio: {producto[4]}, Proveedor: {producto[5]}")

    def registrar_venta(self):
        conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

       # nuevo_id_ventas = self.base_de_datos.generar_id_unico_ventas()

        cursor.execute('SELECT id_producto, producto, cantidad, unidad, proveedor, precio_unitario FROM almacen')  # Selecciona producto, cantidad y unidad
        productos_disponibles = cursor.fetchall()

        if not productos_disponibles:
          print("No hay productos disponibles en el almacén.")
          return

        ventas_realizadas = []  # Lista para almacenar las ventas realizadas
        total_venta = 0  # Inicializar el total de la venta

        while True:
            print("Productos disponibles:")
            for i, (id_producto, producto, cantidad, unidad, proveedor, precio_unitario) in enumerate(productos_disponibles, 1):
                print(f"{i}. ID: {id_producto}, Producto: {producto}, Cantidad: {cantidad} {unidad}, Proveedor: {proveedor}, Precio Unitario: {precio_unitario}")


            producto_idx = int(input("Ingrese el número del producto vendido (0 para finalizar): ")) - 1
            if producto_idx == -1:
                    break

            if 0 <= producto_idx < len(productos_disponibles):
                id_producto, producto, cantidad_disponible, unidad, proveedor, precio_unitario = productos_disponibles[producto_idx]
                cantidad = int(input(f"Ingrese la cantidad vendida (disponible: {cantidad_disponible} {unidad}): "))

                if cantidad <= cantidad_disponible:
                    subtotal = cantidad * precio_unitario  # Calcular el subtotal de la venta
                    total_venta += subtotal  # Actualizar el total de la venta

                    venta = {
                        "producto": producto,
                        "cantidad": cantidad,
                        "precio_unitario": precio_unitario,
                        "unidad": unidad,
                        "proveedor": proveedor,
                        "subtotal": subtotal
                    }

                    ventas_realizadas.append(venta)  # Agregar la venta a la lista
                    nuevo_id_ventas = self.base_de_datos.generar_id_unico_ventas()  # Generar un nuevo ID de venta para cada venta individual

                    cursor.execute('''
                        INSERT INTO ventas (id, producto, cantidad, precio_unitario, unidad, fecha)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (nuevo_id_ventas, venta["producto"], venta["cantidad"], venta["precio_unitario"], venta["unidad"], datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

                    cursor.execute('''
                        UPDATE almacen
                        SET cantidad = cantidad - ?
                        WHERE id_producto = ?
                    ''', (cantidad, id_producto))

                else:
                    print("Número de producto no válido.")
            else:
                print("Número de producto no válido.")

        # Mostrar el resumen de la venta y confirmar
        print("=== Resumen de la Venta ===")
        for venta in ventas_realizadas:
            print(f"Producto: {venta['producto']}, Cantidad: {venta['cantidad']} {venta['unidad']}, Proveedor: {venta['proveedor']}, Precio Unitario: {venta['precio_unitario']}, Subtotal: {venta['subtotal']:.2f}")
        print(f"Total de la Venta: {total_venta:.2f}")

        confirmar_venta = input("¿Desea confirmar la venta? (S/N): ")
        if confirmar_venta.lower() == 's':
            self.exportar_almacen_a_xlsx()
            self.exportar_ventas_a_xlsx()
            conn.commit()
            print("Ventas registradas correctamente.")
        else:
            for venta in ventas_realizadas:
                cursor.execute('''
                    UPDATE almacen
                    SET cantidad = cantidad + ?
                    WHERE id_producto = ?
                ''', (venta["cantidad"], id_producto))
            conn.commit()
            print("Venta cancelada. Se ha devuelto la cantidad de productos al almacén.")

    def ver_reporte_ventas(self):
       # conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT * FROM ventas')
        ventas = cursor.fetchall()

        if not ventas:
            print("No hay ventas registradas.")
        else:
            print("=== Reporte de Ventas ===")
            for venta in ventas:
                print(f"ID: {venta[0]}, Producto: {venta[1]}, Cantidad: {venta[2]}, Unidad: {venta[3]}, Precio: {venta[4]}, Fecha: {venta[5]}")

    def registrar_compra(self):
        while True:
            try:
                proveedor = input("Ingrese el nombre del proveedor (o 'q' para salir): ")
                if proveedor.lower() == 'q':
                    break

                producto = input("Ingrese el nombre del producto comprado: ")
                cantidad = int(input("Ingrese la cantidad comprada: "))
                precio = float(input("Ingrese el precio unitario: "))
                unidad = input("Ingrese la unidad (ejemplo: kilo, litro, paquete, etc.): ")
                fecha = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

                conn = self.base_de_datos.conn
                cursor = self.base_de_datos.cursor

                nuevo_id = self.base_de_datos.generar_id_unico()

                cursor.execute('''
                    INSERT INTO compras (id, proveedor, producto, cantidad, precio_unitario, unidad, fecha)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (nuevo_id, proveedor, producto, cantidad, precio, unidad, fecha))

                cursor.execute('''
                    INSERT OR REPLACE INTO almacen (id_producto, producto, cantidad, unidad, precio_unitario, proveedor)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (nuevo_id, producto, cantidad, unidad, precio, proveedor))

                conn.commit()
                self.exportar_almacen_a_xlsx()
                self.exportar_compras_a_xlsx()

                print("Compra registrada correctamente.")
            except sqlite3.Error as e:
                print("Error al registrar la compra:", e)
            except ValueError:
                print("Por favor, ingrese un valor válido para la cantidad y el precio.")
            except KeyboardInterrupt:
                print("Registro de compra cancelado.")
                break

    def ver_reporte_compras(self):
    #    conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT * FROM compras')
        compras = cursor.fetchall()

        if not compras:
            print("No hay compras registradas.")
        else:
            print("=== Reporte de Compras ===")
            for compra in compras:
                print(f"ID: {compra[0]}, Proveedor: {compra[1]}, Producto: {compra[2]}, Cantidad: {compra[3]}, Precio: {compra[4]}, Unidad: {compra[5]}, Fecha: {compra[6]}")

    def limpiar_contenido_tablas(self):
        try:
            if os.path.exists('ventas.xlsx'):
                os.remove('ventas.xlsx')
            if os.path.exists('compras.xlsx'):
                os.remove('compras.xlsx')
            if os.path.exists('almacen.xlsx'):
                os.remove('almacen.xlsx')
            conn = self.base_de_datos.conn
            cursor = self.base_de_datos.cursor

            cursor.execute('DELETE FROM ventas')
            cursor.execute('DELETE FROM compras')
            cursor.execute('DELETE FROM almacen')

            conn.commit()

            print("El contenido de las tablas 'ventas', 'compras' y 'almacen' y los archivos Excel han sido limpiados.")
        except Exception as e:
            print(f"Ocurrió un error al limpiar el contenido de las tablas: {str(e)}")

    def exportar_almacen_a_xlsx(self):
       # conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT * FROM almacen')
        almacen = cursor.fetchall()

        if not almacen:
            print("No hay datos en el almacén para exportar.")
            return

        # Crear un nuevo libro de Excel
        workbook = Workbook()
        sheet = workbook.active

        # Agregar encabezados de columna
        sheet.append(['id_producto', 'Producto', 'Cantidad', 'Unidad', 'Precio', 'Proveedor'])

        # Agregar datos
        for row in almacen:
            sheet.append(row)

        # Guardar el libro de Excel como archivo XLSX
        workbook.save('almacen.xlsx')

        print("Datos del almacén exportados a almacen.xlsx")

    def exportar_ventas_a_xlsx(self):
        # Obtener la fecha actual en formato YYYY-MM-DD
        fecha_actual = datetime.now().strftime('%Y-%m-%d')

        # Crear un nuevo archivo Excel
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = fecha_actual  # Usar la fecha actual como título de la hoja

        # Agregar las ventas registradas en el día
        cursor = self.base_de_datos.cursor
        cursor.execute('SELECT producto, cantidad, precio_unitario FROM ventas WHERE fecha LIKE ?', (f'{fecha_actual}%',))
        ventas_del_dia = cursor.fetchall()

        # Agregar las ventas al archivo Excel
        sheet['A1'] = 'Producto'
        sheet['B1'] = 'Cantidad'
        sheet['C1'] = 'Precio Unitario'
        sheet['D1'] = 'Subtotal'

        total_ventas_del_dia = 0  # Inicializar el total de ventas del día

        for i, (producto, cantidad, precio_unitario) in enumerate(ventas_del_dia, 2):
            sheet[f'A{i}'] = producto
            sheet[f'B{i}'] = cantidad
            sheet[f'C{i}'] = precio_unitario
            subtotal = cantidad * precio_unitario
            sheet[f'D{i}'] = subtotal
            total_ventas_del_dia += subtotal

        # Calcular y agregar la suma total de ventas del día
        total_row = len(ventas_del_dia) + 3  # Separación de dos celdas y una fila adicional
        sheet[f'A{total_row}'] = 'Total del Día:'
        sheet[f'D{total_row}'] = total_ventas_del_dia

        # Establecer el formato de la celda para el total
        total_cell = sheet[f'D{total_row}']
        total_cell.font = Font(bold=True)

        # Guardar el archivo Excel
        archivo_xlsx = f'ventas_{fecha_actual}.xlsx'
        workbook.save(archivo_xlsx)

        print(f"Reporte de ventas del día {fecha_actual} exportado a {archivo_xlsx}")

    def exportar_compras_a_xlsx(self):
     #   conn = self.base_de_datos.conn
        cursor = self.base_de_datos.cursor

        cursor.execute('SELECT * FROM compras')
        compras = cursor.fetchall()

        if not compras:
            print("No hay datos de compras para exportar.")
            return

        # Crear un nuevo libro de Excel
        workbook = Workbook()
        sheet = workbook.active

        # Agregar encabezados de columna
        sheet.append(['ID', 'Proveedor', 'Producto', 'Cantidad', 'Unidad', 'Precio', 'Fecha'])

        # Agregar datos
        for row in compras:
            sheet.append(row)

        # Guardar el libro de Excel como archivo XLSX
        workbook.save('compras.xlsx')

        print("Datos de compras exportados a compras.xlsx")
        
    def ver_reporte_compras_almacen(self):
      # conn = self.base_de_datos.conn
       cursor = self.base_de_datos.cursor
       
       cursor.execute('SELECT producto, unidad, precio_unitario, SUM(cantidad) AS total_cantidad, SUM(cantidad * precio_unitario) AS total_costo FROM compras GROUP BY producto, unidad')
       compras_por_producto = cursor.fetchall()

       if not compras_por_producto:
         print("No hay compras registradas en el almacén.")
         return

       print("=== Reporte de Compras para el Almacén ===")
       for producto, unidad, precio_unitario, cantidad_total, total_costo in compras_por_producto:
           print(f"Producto: {producto} ({unidad}), Precio_Unitario: {precio_unitario}, Cantidad: {cantidad_total}, Costo Total: {total_costo}")
           
           suma_total_costos = sum(total_costo for _, _, precio_unitario, cantidad_total, total_costo in compras_por_producto)
       print(f"Suma Total de Costos: {suma_total_costos}")

       # Exporta el reporte a un archivo Excel
       self.exportar_reporte_compras_almacen_a_xlsx(compras_por_producto, suma_total_costos)

    def exportar_reporte_compras_almacen_a_xlsx(self, compras_por_producto, suma_total_costos):
       workbook = Workbook()
       sheet = workbook.active
       sheet.title = "Reporte Compras Almacén"

       # Agregar encabezados de columna
       sheet.append(["Producto", "Unidad", "Precio Unitario", "Cantidad", "Sub-Suma", "Costo Total"])

       # Agregar datos
       for producto, unidad, precio_unitario, total_cantidad, total_costo in compras_por_producto:
           sub_suma = total_costo  # Sub-suma es igual al costo total en este caso
           sheet.append([producto, unidad, precio_unitario, total_cantidad, sub_suma, total_costo])

       # Agregar fila con la suma total de costos
       sheet.append(["", "", "", "", "Suma Total de Costos", suma_total_costos])

       # Guardar el libro de Excel como archivo XLSX
       workbook.save('reporte_compras_almacen.xlsx')

       print("Reporte de Compras para el Almacén exportado a reporte_compras_almacen.xlsx")    

    def modificar_precio_producto(self):
        cursor = self.base_de_datos.cursor

        # Obtener la lista de productos disponibles en el almacén junto con sus atributos
        cursor.execute('SELECT id_producto, producto, cantidad, unidad, proveedor, precio_unitario FROM almacen')
        productos_disponibles = cursor.fetchall()

        if not productos_disponibles:
            print("No hay productos disponibles en el almacén.")
            return

        print("Productos disponibles:")
        for i, (id_producto, producto, cantidad, unidad, proveedor, precio_unitario) in enumerate(productos_disponibles, 1):
            print(f"{i}. ID: {id_producto}, Producto: {producto}, Cantidad: {cantidad} {unidad}, Proveedor: {proveedor}, Precio: {precio_unitario}")

        try:
            producto_idx = int(input("Ingrese el número del producto cuyo precio desea modificar: ")) - 1
            if 0 <= producto_idx < len(productos_disponibles):
                id_producto, _, _, _, _, _ = productos_disponibles[producto_idx]
                nuevo_precio = float(input("Ingrese el nuevo precio unitario: "))

                # Actualizar el precio unitario en la tabla 'almacen'
                cursor.execute('''
                    UPDATE almacen
                    SET precio_unitario = ?
                    WHERE id_producto = ?
                ''', (nuevo_precio, id_producto))

                self.base_de_datos.conn.commit()
                print(f"Precio del producto con ID '{id_producto}' actualizado correctamente.")
            else:
                print("Número de producto no válido.")
        except ValueError:
            print("Por favor, ingrese un valor válido para el precio.")


   
if __name__ == '__main__':
    logica = LogicaDeNegocios()
    logica.base_de_datos.inicializar_base_de_datos()
    logica.menu_principal()